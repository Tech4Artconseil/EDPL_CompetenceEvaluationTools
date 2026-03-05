"""
App_Main.py
Flask application with routes and logic for student competence evaluation.
Handles dashboard display, score updates, comments, and CSV export.
"""

import os
import sys
import webbrowser
import threading
from flask import Flask, render_template, request, jsonify, send_file
import re
from sqlalchemy import nullslast
from Data_Models import Db, Level, Skill, Studnt, Evaluat, Score, Comment, Note, EvaluatNote, Saison, EvaluatGrp
from admin import admin_bp
import csv
import io
from datetime import datetime
from sheets_local import preview_mappings, fill_template, fill_student_sheets
from Data_Models import MappingType
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# --- Détection mode PyInstaller (frozen) ou développement ---
# En mode frozen : les fichiers extraits (templates/static) sont dans sys._MEIPASS
# En mode développement : tout est dans le répertoire du script
if getattr(sys, 'frozen', False):
    # Exécutable compilé par PyInstaller
    BASE_DIR = os.path.dirname(sys.executable)   # dossier contenant le .exe / binaire
    BUNDLE_DIR = sys._MEIPASS                     # fichiers embarqués extraits au lancement
else:
    BASE_DIR = os.path.abspath(os.path.dirname(__file__))
    BUNDLE_DIR = BASE_DIR

# Initialize Flask application with explicit paths (compatibles frozen + dev)
App = Flask(
    __name__,
    template_folder=os.path.join(BUNDLE_DIR, 'templates'),
    static_folder=os.path.join(BUNDLE_DIR, 'static'),
)
# La base SQLite est stockée à côté de l'exécutable (persistance entre sessions)
PACKAGE_DIR = BASE_DIR
INSTANCE_DIR = os.path.join(BASE_DIR, 'instance')
os.makedirs(INSTANCE_DIR, exist_ok=True)
DB_PATH = os.path.join(INSTANCE_DIR, 'evaluat.db')
# Use forward slashes in the URI to avoid Windows backslash issues
App.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + DB_PATH.replace('\\', '/')
App.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
App.config['SECRET_KEY'] = 'dev-secret-key-change-in-production'

# Initialize database with app
Db.init_app(App)

# Register admin blueprint
App.register_blueprint(admin_bp)
from import_csv import importer_bp
App.register_blueprint(importer_bp)


def norm_color(val: object) -> str:
    """Normalize a color value for CSS use. Returns a hex color like '#rrggbb' or the original string.

    Accepts values already in '#rrggbb' or 'rrggbb' or common color names (red, green...).
    """
    if not val:
        return ''
    s = str(val).strip()
    sl = s.lower()
    # common name map
    name_map = {
        'red': '#ff0000', 'green': '#00ff00', 'blue': '#0000ff', 'yellow': '#ffff00',
        'white': '#ffffff', 'black': '#000000', 'grey': '#808080', 'gray': '#808080'
    }
    if sl in name_map:
        return name_map[sl]
    # already with #
    if sl.startswith('#'):
        if re.match(r'^#([0-9a-f]{3}|[0-9a-f]{6})$', sl):
            if len(sl) == 4:  # expand #rgb to #rrggbb
                return '#' + sl[1]*2 + sl[2]*2 + sl[3]*2
            return sl
        return sl
    # hex without #
    if re.match(r'^[0-9a-f]{3}$', sl):
        return '#' + ''.join(ch*2 for ch in sl)
    if re.match(r'^[0-9a-f]{6}$', sl):
        return '#' + sl
    return s


@App.route('/')
def App_Main_Dashboard():
    """
    Main dashboard route - displays evaluation grid.
    Shows students (rows) vs skills (columns) with level selection.
    """
    # Allow selecting an evaluation via query param ?evaluat_id=ID
    evaluat_id = request.args.get('evaluat_id', type=int)
    # Allow filtering the panel list by saison via ?saison=2025-2026
    saison_filter = request.args.get('saison', type=str) or None

    Current_Evaluat = None
    if evaluat_id:
        Current_Evaluat = Evaluat.query.get(evaluat_id)

    # If none selected, pick the first evaluation of the filtered saison (or globally first)
    if not Current_Evaluat:
        q = Evaluat.query
        if saison_filter:
            q = q.filter(Evaluat.Saison_Rel.has(Name=saison_filter))
        Current_Evaluat = q.order_by(Evaluat.CreatedAt.desc()).first()

    if not Current_Evaluat:
        Current_Evaluat = Evaluat.query.order_by(Evaluat.CreatedAt.desc()).first()

    if not Current_Evaluat:
        return "No evaluation found. Please run Data_Init.py first.", 404

    # Collect all saisons for the filter bar (from the saisons table)
    All_Saisons = sorted([s.Name for s in Saison.query.all()], reverse=True)

    # Load all evaluations for the panels, grouped by saison then name
    Evaluat_Query = Evaluat.query.outerjoin(Saison, Evaluat.Saison_Id == Saison.Id)
    if saison_filter:
        Evaluat_Query = Evaluat_Query.filter(Saison.Name == saison_filter)
    All_Evaluats = Evaluat_Query.order_by(
        nullslast(Saison.Name.desc()),
        Evaluat.CreatedAt.desc()
    ).all()
    
    # Get all students in the evaluation's group
    All_Studnts = Studnt.query.filter_by(Group_Id=Current_Evaluat.Group_Id).order_by(Studnt.Name).all()
    
    # Get all skills in the evaluation's skill set
    All_Skills = Skill.query.filter_by(SkillSet_Id=Current_Evaluat.SkillSet_Id).order_by(Skill.Code).all()
    
    # Get all levels for the level selector
    All_Levels = Level.query.filter_by(LevelSet_Id=1).order_by(Level.Percent).all()
    
    # Get all scores for this evaluation
    # Create a dictionary for quick lookup: (student_id, skill_id) -> score
    All_Scores = Score.query.filter_by(Evaluat_Id=Current_Evaluat.Id).all()
    Scores_Dict = {}
    for Score_tmp in All_Scores:
        Key = f"{Score_tmp.Studnt_Id}_{Score_tmp.Skill_Id}"
        Scores_Dict[Key] = Score_tmp
    
    # Get all comments for this evaluation
    All_Comments = Comment.query.filter_by(Evaluat_Id=Current_Evaluat.Id).all()
    Comments_Dict = {}
    for Comment_tmp in All_Comments:
        Key = f"{Comment_tmp.Studnt_Id}_{Comment_tmp.Skill_Id}"
        if Key not in Comments_Dict:
            Comments_Dict[Key] = []
        Comments_Dict[Key].append(Comment_tmp)

    # Build a mapping of student_id -> comments for comments that are NOT tied to a skill
    Student_Comments_Dict = {}
    for Comment_tmp in All_Comments:
        if Comment_tmp.Skill_Id is None:
            sid = Comment_tmp.Studnt_Id
            if sid not in Student_Comments_Dict:
                Student_Comments_Dict[sid] = []
            Student_Comments_Dict[sid].append(Comment_tmp)

    # Get all notes and build mapping student_id -> note_id for this evaluation
    All_Notes = Note.query.order_by(Note.Valeure.desc()).all()
    Student_Notes_Dict = {}
    All_Evaluat_Notes = EvaluatNote.query.filter_by(Evaluat_Id=Current_Evaluat.Id).all()
    for en in All_Evaluat_Notes:
        Student_Notes_Dict[en.Studnt_Id] = en.Note_Id

    # Load sous-groupes for this evaluation (sorted by name for consistent display)
    Evaluat_Grps = EvaluatGrp.query.filter_by(Evaluat_Id=Current_Evaluat.Id).order_by(EvaluatGrp.Name).all()
    # Build set of student IDs that appear in at least one sous-groupe
    Grp_Studnt_Ids = set()
    for _grp in Evaluat_Grps:
        for _m in _grp.Members:
            Grp_Studnt_Ids.add(_m.Studnt_Id)

    return render_template(
        'Eval_Dash.html',
        Evaluat=Current_Evaluat,
        All_Evaluats=All_Evaluats,
        All_Saisons=All_Saisons,
        Saison_Filter=saison_filter,
        Studnts=All_Studnts,
        Skills=All_Skills,
        Levels=All_Levels,
        All_Notes=All_Notes,
        Scores_Dict=Scores_Dict,
        Comments_Dict=Comments_Dict,
        Student_Comments_Dict=Student_Comments_Dict,
        Student_Notes_Dict=Student_Notes_Dict,
        Evaluat_Grps=Evaluat_Grps,
        Grp_Studnt_Ids=Grp_Studnt_Ids,
        norm_color=norm_color
    )


@App.route('/evaluat/sheet/preview')
def Evaluat_Sheet_Preview():
    evaluat_id = request.args.get('evaluat_id', type=int)
    mapping_type_id = request.args.get('mapping_type_id', type=int)
    if not evaluat_id:
        return "evaluat_id required", 400
    # preview either mapping type or evaluation-linked mappings (default: mapping_type if provided)
    data = preview_mappings(evaluat_id=evaluat_id, mapping_type_id=mapping_type_id) if mapping_type_id else preview_mappings(evaluat_id=evaluat_id)
    if 'error' in data:
        return data['error'], 404
    # pass available mapping types for selection
    mapping_types = MappingType.query.order_by(MappingType.Name).all()
    return render_template('sheet_preview.html', preview=data, mapping_types=mapping_types, selected_mapping_type=mapping_type_id)


@App.route('/evaluat/sheet/fill', methods=['POST'])
def Evaluat_Sheet_Fill():
    evaluat_id = request.form.get('evaluat_id', type=int)
    mapping_type_id = request.form.get('mapping_type_id', type=int)
    if not evaluat_id:
        return "evaluat_id required", 400
    result = fill_template(evaluat_id=evaluat_id, mapping_type_id=mapping_type_id)
    if not result.get('success'):
        return result.get('error', 'Unknown error'), 500
    out_path = result.get('path')
    if not out_path or not os.path.exists(out_path):
        return "Output file not found", 500
    # send the file as attachment
    return send_file(out_path, as_attachment=True, download_name=os.path.basename(out_path))


@App.route('/evaluat/sheet/fill_simple', methods=['POST'])
def Evaluat_Sheet_Fill_Simple():
    evaluat_id = request.form.get('evaluat_id', type=int)
    dry_run = request.form.get('dry_run', '0') == '1'
    if not evaluat_id:
        return "evaluat_id required", 400
    result = fill_student_sheets(evaluat_id=evaluat_id, dry_run=dry_run)
    if not result.get('success'):
        return result.get('error', 'Unknown error'), 500
    if dry_run:
        return jsonify(result)
    out_path = result.get('path')
    if not out_path or not os.path.exists(out_path):
        return "Output file not found", 500
    return send_file(out_path, as_attachment=True, download_name=os.path.basename(out_path))



@App.route('/api/evaluat/note/set', methods=['POST'])
def App_Main_Evaluat_Note_Set():
    """Assign or remove a Note for a student in an evaluation.
    JSON: {evaluat_id, studnt_id, note_id|null}
    """
    Data = request.get_json() or {}
    evaluat_id = Data.get('evaluat_id')
    studnt_id = Data.get('studnt_id')
    note_id = Data.get('note_id')  # can be null to remove

    if not evaluat_id or not studnt_id:
        return jsonify({'success': False, 'error': 'evaluat_id and studnt_id required'}), 400

    try:
        existing = EvaluatNote.query.filter_by(Evaluat_Id=evaluat_id, Studnt_Id=studnt_id).first()
        if not note_id:
            if existing:
                Db.session.delete(existing)
                Db.session.commit()
            return jsonify({'success': True})

        note = Note.query.get(int(note_id))
        if not note:
            return jsonify({'success': False, 'error': 'Note not found'}), 404

        if existing:
            existing.Note_Id = note.Id
        else:
            en = EvaluatNote(Evaluat_Id=evaluat_id, Studnt_Id=studnt_id, Note_Id=note.Id)
            Db.session.add(en)
        Db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        Db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@App.route('/api/score/update', methods=['POST'])
def App_Main_Score_Update():
    """
    API endpoint to update a student's score for a skill.
    Receives JSON with student_id, skill_id, and level_id.
    """
    Data = request.get_json()
    
    Studnt_Id = Data.get('studnt_id')
    Skill_Id = Data.get('skill_id')
    Level_Id = Data.get('level_id')
    Evaluat_Id = Data.get('evaluat_id')
    
    # Find the score entry
    Score_Entry = Score.query.filter_by(
        Evaluat_Id=Evaluat_Id,
        Studnt_Id=Studnt_Id,
        Skill_Id=Skill_Id
    ).first()
    
    if not Score_Entry:
        return jsonify({'success': False, 'error': 'Score entry not found'}), 404
    
    # Update the level
    Score_Entry.Level_Id = Level_Id if Level_Id else None
    Db.session.commit()
    
    return jsonify({'success': True})


@App.route('/api/notes/list', methods=['GET'])
def App_Main_Notes_List():
    """Return list of all notes as JSON."""
    All_Notes = Note.query.order_by(Note.Valeure.desc()).all()
    notes = [n.Note_To_Dict() for n in All_Notes]
    return jsonify({'success': True, 'notes': notes})


@App.route('/api/notes/add', methods=['POST'])
def App_Main_Notes_Add():
    """Create or update a Note. JSON body: {id?, valeure, descript}
    If id provided, update; otherwise create new.
    """
    Data = request.get_json() or {}
    nid = Data.get('id')
    valeure = Data.get('valeure')
    descript = Data.get('descript')

    if valeure is None:
        return jsonify({'success': False, 'error': 'Field "valeure" is required'}), 400

    try:
        if nid:
            note = Note.query.get(int(nid))
            if not note:
                return jsonify({'success': False, 'error': 'Note not found'}), 404
            note.Valeure = int(valeure)
            note.Descript = descript
        else:
            note = Note(Valeure=int(valeure), Descript=descript)
            Db.session.add(note)
        Db.session.commit()
        return jsonify({'success': True, 'note': note.Note_To_Dict()})
    except Exception as e:
        Db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@App.route('/api/notes/delete', methods=['POST'])
def App_Main_Notes_Delete():
    """Delete a note. JSON body: {id}
    """
    Data = request.get_json() or {}
    nid = Data.get('id')
    if not nid:
        return jsonify({'success': False, 'error': 'Field "id" is required'}), 400
    try:
        note = Note.query.get(int(nid))
        if not note:
            return jsonify({'success': False, 'error': 'Note not found'}), 404
        Db.session.delete(note)
        Db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        Db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@App.route('/api/comment/add', methods=['POST'])
def App_Main_Comment_Add():
    """
    API endpoint to add a comment for a student/skill combination.
    Receives JSON with student_id, skill_id, and comment text.
    """
    Data = request.get_json()
    
    Studnt_Id = Data.get('studnt_id')
    Skill_Id = Data.get('skill_id')
    Comment_Text = Data.get('comment_text')
    Evaluat_Id = Data.get('evaluat_id')
    
    if not Comment_Text or not Comment_Text.strip():
        return jsonify({'success': False, 'error': 'Comment text is required'}), 400
    
    # Create new comment
    try:
        New_Comment = Comment(
            Evaluat_Id=Evaluat_Id,
            Studnt_Id=Studnt_Id,
            Skill_Id=Skill_Id,
            Text=Comment_Text.strip()
        )
        Db.session.add(New_Comment)
        Db.session.commit()
    except Exception as e:
        # Rollback and return JSON error so client-side JSON.parse won't fail
        Db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

    return jsonify({
        'success': True,
        'comment': New_Comment.Comment_To_Dict()
    })


@App.route('/api/comment/list', methods=['GET'])
def App_Main_Comment_List():
    """
    API endpoint to get comments for a student/skill combination.
    Query params: studnt_id, skill_id, evaluat_id
    """
    Studnt_Id = request.args.get('studnt_id', type=int)
    Skill_Id = request.args.get('skill_id', type=int)
    Evaluat_Id = request.args.get('evaluat_id', type=int)
    
    # Get all comments for this combination
    All_Comments = Comment.query.filter_by(
        Evaluat_Id=Evaluat_Id,
        Studnt_Id=Studnt_Id,
        Skill_Id=Skill_Id
    ).order_by(Comment.CreatedAt.desc()).all()
    
    Comments_List = [Comment_tmp.Comment_To_Dict() for Comment_tmp in All_Comments]
    
    return jsonify({'success': True, 'comments': Comments_List})


@App.route('/export/csv')
def App_Main_Export_CSV():
    """
    Export evaluation data to CSV file with semicolon separator.
    Creates a CSV with students as rows and skills as columns.
    """
    # Allow exporting a specific evaluation via query param ?evaluat_id=ID
    evaluat_id = request.args.get('evaluat_id', type=int)
    if evaluat_id:
        Current_Evaluat = Evaluat.query.get(evaluat_id)
    else:
        Current_Evaluat = Evaluat.query.first()

    if not Current_Evaluat:
        return "No evaluation found", 404
    
    # Get all students and skills
    All_Studnts = Studnt.query.filter_by(Group_Id=Current_Evaluat.Group_Id).order_by(Studnt.Name).all()
    All_Skills = Skill.query.filter_by(SkillSet_Id=Current_Evaluat.SkillSet_Id).order_by(Skill.Code).all()
    
    # Get all scores
    All_Scores = Score.query.filter_by(Evaluat_Id=Current_Evaluat.Id).all()
    Scores_Dict = {}
    for Score_tmp in All_Scores:
        Key = (Score_tmp.Studnt_Id, Score_tmp.Skill_Id)
        Scores_Dict[Key] = Score_tmp

    # Get all notes assigned for this evaluation (per-student)
    All_Evaluat_Notes = EvaluatNote.query.filter_by(Evaluat_Id=Current_Evaluat.Id).all()
    Student_Note_Map = {en.Studnt_Id: en.Note_Id for en in All_Evaluat_Notes}

    # Load notes lookup
    All_Notes = {n.Id: n for n in Note.query.all()}

    # Load comments for this evaluation and aggregate per-student
    All_Comments = Comment.query.filter_by(Evaluat_Id=Current_Evaluat.Id).all()
    # build skill id -> code map for nicer comment labels
    SkillCode = {s.Id: s.Code for s in All_Skills}
    Student_Comments_Map = {}
    for c in All_Comments:
        sid = c.Studnt_Id
        skill_label = SkillCode.get(c.Skill_Id) if c.Skill_Id else 'General'
        text = (c.Text or '').replace('\n', ' ').replace(';', ',').strip()
        entry = f"[{skill_label}] {text}"
        Student_Comments_Map.setdefault(sid, []).append(entry)
    
    # Create CSV in memory
    Output = io.StringIO()
    Writer = csv.writer(Output, delimiter=';')
    
    # Write header row (include Note and aggregated Comments) - email removed
    Header = ['Student']
    for Skill_tmp in All_Skills:
        Header.append(f'{Skill_tmp.Code}')
    Header.append('Note')
    Header.append('Comments')
    Writer.writerow(Header)
    
    # Write data rows - one row per student
    for Studnt_tmp in All_Studnts:
        Row = [Studnt_tmp.Name]
        
        # Add score for each skill
        for Skill_tmp in All_Skills:
            Key = (Studnt_tmp.Id, Skill_tmp.Id)
            Score_Entry = Scores_Dict.get(Key)
            
            if Score_Entry and Score_Entry.Level:
                # Write level description (use Descrip, not percent)
                Row.append(f'{Score_Entry.Level.Descrip}')
            else:
                # No level assigned
                Row.append('')
        
        # Add Note value (if any)
        note_val = ''
        nid = Student_Note_Map.get(Studnt_tmp.Id)
        if nid:
            note = All_Notes.get(nid)
            if note:
                # Only include the numeric value for exports (no description)
                note_val = f"{note.Valeure}"
        Row.append(note_val)

        # Add aggregated comments (joined by ' | ')
        comments_list = Student_Comments_Map.get(Studnt_tmp.Id, [])
        comments_cell = ' | '.join(comments_list)
        Row.append(comments_cell)

        Writer.writerow(Row)
    
    # Prepare file for download
    Output.seek(0)
    Output_Bytes = io.BytesIO()
    Output_Bytes.write(Output.getvalue().encode('utf-8'))
    Output_Bytes.seek(0)
    
    # Generate filename with timestamp
    Filename = f'evaluat_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    
    return send_file(
        Output_Bytes,
        mimetype='text/csv',
        as_attachment=True,
        download_name=Filename
    )


@App.route('/export/xlsx')
def App_Main_Export_XLSX():
    """Export evaluation data to XLSX with colored cells for levels.
    Falls back to a 400 error if openpyxl is not installed.
    """
    if not OPENPYXL_AVAILABLE:
        return "openpyxl not installed. Install requirements and restart.", 400

    evaluat_id = request.args.get('evaluat_id', type=int)
    if evaluat_id:
        Current_Evaluat = Evaluat.query.get(evaluat_id)
    else:
        Current_Evaluat = Evaluat.query.first()

    if not Current_Evaluat:
        return "No evaluation found", 404

    All_Studnts = Studnt.query.filter_by(Group_Id=Current_Evaluat.Group_Id).order_by(Studnt.Name).all()
    All_Skills = Skill.query.filter_by(SkillSet_Id=Current_Evaluat.SkillSet_Id).order_by(Skill.Code).all()
    All_Scores = Score.query.filter_by(Evaluat_Id=Current_Evaluat.Id).all()
    Scores_Dict = {(s.Studnt_Id, s.Skill_Id): s for s in All_Scores}

    # Notes and comments mapping (reuse logic from CSV exporter)
    All_Evaluat_Notes = EvaluatNote.query.filter_by(Evaluat_Id=Current_Evaluat.Id).all()
    Student_Note_Map = {en.Studnt_Id: en.Note_Id for en in All_Evaluat_Notes}
    All_Notes = {n.Id: n for n in Note.query.all()}

    All_Comments = Comment.query.filter_by(Evaluat_Id=Current_Evaluat.Id).all()
    SkillCode = {s.Id: s.Code for s in All_Skills}
    Student_Comments_Map = {}
    for c in All_Comments:
        sid = c.Studnt_Id
        skill_label = SkillCode.get(c.Skill_Id) if c.Skill_Id else 'General'
        text = (c.Text or '').replace('\n', ' ').strip()
        entry = f"[{skill_label}] {text}"
        Student_Comments_Map.setdefault(sid, []).append(entry)

    wb = Workbook()
    ws = wb.active
    ws.title = Current_Evaluat.Name[:31]

    # Header (Email removed)
    headers = ['Student'] + [s.Code for s in All_Skills] + ['Note', 'Comments']
    ws.append(headers)

    # helper: convert hex color to openpyxl PatternFill
    def fill_from_color(color_str):
        if not color_str:
            return None
        s = color_str.strip()
        if s.startswith('#'):
            hexc = s[1:]
        else:
            hexc = s
        if len(hexc) == 3:
            hexc = ''.join(ch*2 for ch in hexc)
        if len(hexc) != 6:
            return None
        return PatternFill(start_color=hexc.upper(), end_color=hexc.upper(), fill_type='solid')

    # Write rows
    for st in All_Studnts:
        row = [st.Name]
        for sk in All_Skills:
            sc = Scores_Dict.get((st.Id, sk.Id))
            if sc and sc.Level:
                # Use level description instead of percent
                val = f"{sc.Level.Descrip}"
            else:
                val = ''
            row.append(val)
        # Note
        nid = Student_Note_Map.get(st.Id)
        note_text = ''
        if nid:
            n = All_Notes.get(nid)
            if n:
                # Only include the numeric value for exports (no description)
                note_text = str(n.Valeure)
        row.append(note_text)
        # Comments
        comments_cell = ' | '.join(Student_Comments_Map.get(st.Id, []))
        row.append(comments_cell)
        ws.append(row)

    # Apply fills to skill cells based on Level colors
    # header offset: Student=1, skills start at col 2 (Email removed)
    for r_idx, st in enumerate(All_Studnts, start=2):
        for c_idx, sk in enumerate(All_Skills, start=2):
            sc = Scores_Dict.get((st.Id, sk.Id))
            if sc and sc.Level and sc.Level.Color:
                fill = fill_from_color(sc.Level.Color)
                if fill:
                    ws.cell(row=r_idx, column=c_idx).fill = fill

    # Prepare output
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    fname = f'evaluat_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=fname)


if __name__ == '__main__':
    # Créer les tables si elles n'existent pas encore (première utilisation sur une nouvelle machine)
    with App.app_context():
        Db.create_all()

    if getattr(sys, 'frozen', False):
        # ---- Mode application empaquetée (PyInstaller) ----
        # Ouvrir automatiquement le navigateur après démarrage du serveur
        def _open_browser():
            import time
            time.sleep(1.5)  # laisser Flask démarrer
            webbrowser.open('http://localhost:5000')
        threading.Thread(target=_open_browser, daemon=True).start()
        print("EDPL Competence Evaluation Tool")
        print("Serveur démarré sur http://localhost:5000")
        print("Fermez cette fenêtre pour arrêter l'application.")
        App.run(debug=False, host='127.0.0.1', port=5000)
    else:
        # ---- Mode développement ----
        App.run(debug=True, host='0.0.0.0', port=5000)
