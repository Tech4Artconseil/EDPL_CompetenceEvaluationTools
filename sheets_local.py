import os
import json
from datetime import datetime
from openpyxl import load_workbook
from Data_Models import Db, Evaluat, Studnt, Skill, Note, Score, Comment, SheetMapping, MappingType
from sqlalchemy.exc import NoResultFound
from sqlalchemy import func
import re
from sqlalchemy import func


def _split_formula_args(s: str):
    """Split formula arguments on comma or semicolon, respecting simple cases."""
    # choose separator: prefer ';' if present (French Excel), else comma
    sep = ';' if ';' in s else ','
    parts = [p.strip() for p in s.split(sep)]
    return parts


def _eval_vlookup_formula(formula: str, wb, wb_data, current_sheet: str):
    """Try to evaluate a simple VLOOKUP/RECHERCHEV formula that points to a table in the workbook.
    Returns the found value or None. This supports formulas like:
      =RECHERCHEV(A3;'Groupe et bilan'!A13:E47;2)
      =VLOOKUP(A3,'Groupe et bilan'!A13:E47,2,FALSE)
    It does a basic exact-match search on the first column of the table.
    """
    if not formula or not isinstance(formula, str):
        return None
    f = formula.strip()
    if f.startswith('='):
        f = f[1:]
    # normalize function name
    up = f.upper()
    if 'RECHERCHEV' not in up and 'VLOOKUP' not in up:
        return None
    # extract arguments inside parentheses
    try:
        start = f.index('(')
        end = f.rindex(')')
        args_str = f[start+1:end]
    except Exception:
        return None
    args = _split_formula_args(args_str)
    if len(args) < 3:
        return None

    lookup_ref = args[0]
    table_ref = args[1]
    try:
        col_idx = int(args[2])
    except Exception:
        return None

    # resolve lookup value (lookup_ref may be like A3 or Sheet!A3)
    def resolve_cell(ref, use_data=True, base_sheet=current_sheet):
        ref = ref.strip()
        if '!' in ref:
            sh, cref = ref.split('!', 1)
            sh = sh.strip().strip("'")
            cref = cref.strip()
        else:
            sh = base_sheet
            cref = ref
        wb_src = wb_data if use_data else wb
        if sh not in wb_src.sheetnames:
            return None
        try:
            return wb_src[sh][cref].value
        except Exception:
            return None

    lookup_val = resolve_cell(lookup_ref, use_data=True)
    if lookup_val is None:
        # try evaluating from non-data workbook
        lookup_val = resolve_cell(lookup_ref, use_data=False)
    if lookup_val is None:
        return None

    # parse table ref: may be like 'Sheet'!A13:E47 or Sheet!A13:E47
    tr = table_ref.strip()
    if '!' in tr:
        sh, rng = tr.split('!', 1)
        sh = sh.strip().strip("'")
        rng = rng.strip()
    else:
        # no sheet specified; use current sheet
        sh = current_sheet
        rng = tr

    if sh not in wb_data.sheetnames:
        return None

    # rng expected like A13:E47
    if ':' not in rng:
        return None
    start_cell, end_cell = [c.strip() for c in rng.split(':', 1)]

    try:
        start_col = ''.join([ch for ch in start_cell if ch.isalpha()])
        start_row = int(''.join([ch for ch in start_cell if ch.isdigit()]))
        end_col = ''.join([ch for ch in end_cell if ch.isalpha()])
        end_row = int(''.join([ch for ch in end_cell if ch.isdigit()]))
    except Exception:
        return None

    ws_tbl = wb_data[sh]
    # iterate rows and match first column
    for r in range(start_row, end_row+1):
        try:
            cell_val = ws_tbl[f"{start_col}{r}"].value
        except Exception:
            cell_val = None
        if cell_val is None:
            continue
        if str(cell_val).strip() == str(lookup_val).strip():
            # found: return value from column col_idx
            # compute target column letter: start_col + (col_idx-1) offset
            # convert column letters to index
            def col_to_index(col):
                idx = 0
                for ch in col.upper():
                    idx = idx*26 + (ord(ch)-64)
                return idx
            def index_to_col(idx):
                s = ''
                while idx>0:
                    idx, r = divmod(idx-1, 26)
                    s = chr(65+r) + s
                return s

            start_idx = col_to_index(start_col)
            target_idx = start_idx + (col_idx-1)
            target_col = index_to_col(target_idx)
            try:
                return wb_data[sh][f"{target_col}{r}"].value
            except Exception:
                return None

    return None


def _normalize_skill_code(s: str):
    if s is None:
        return ''
    # keep alphanum characters only, uppercase
    return ''.join(ch for ch in str(s).upper() if ch.isalnum())


def _eval_indirect_formula(formula: str, wb, wb_data, current_sheet: str):
    """Evaluate a simple INDIRECT pattern like INDIRECT($C$3&"!E6") inside formulas.
    Returns the target cell value or None.
    """
    if not formula or not isinstance(formula, str):
        return None
    f = formula.strip()
    if f.startswith('='):
        f = f[1:]
    if 'INDIRECT' not in f.upper():
        return None
    # find the first cell ref used as sheet-name source (e.g. $C$3)
    m = re.search(r"\$?[A-Za-z]+\$?\d+", f)
    if not m:
        return None
    sheet_ref = m.group(0)

    # find the target cell after a '!'
    m2 = re.search(r"!([A-Z]+\d+)", f)
    if not m2:
        return None
    target_cell = m2.group(1)

    # resolve sheet name from sheet_ref
    def resolve_cell_value(ref, use_data=True, base_sheet=current_sheet):
        ref = ref.strip()
        if '!' in ref:
            sh, cref = ref.split('!', 1)
            sh = sh.strip().strip("'")
            cref = cref.strip()
        else:
            sh = base_sheet
            cref = ref
        wb_src = wb_data if use_data else wb
        if sh not in wb_src.sheetnames:
            return None
        try:
            return wb_src[sh][cref].value
        except Exception:
            return None

    sheet_name_val = resolve_cell_value(sheet_ref, use_data=True)
    if sheet_name_val is None:
        sheet_name_val = resolve_cell_value(sheet_ref, use_data=False)
    if sheet_name_val is None:
        return None

    sheet_name = str(sheet_name_val).strip().strip("'")
    if sheet_name not in wb_data.sheetnames:
        return None

    try:
        return wb_data[sheet_name][target_cell].value
    except Exception:
        return None


# Map table names (as stored in Source_Table) to model classes
TABLE_MODEL_MAP = {
    'evaluats': Evaluat,
    'studnts': Studnt,
    'skills': Skill,
    'notes': Note,
    'scores': Score,
    'comments': Comment,
}


def _parse_simple_filter(filter_str: str):
    """Parse a very small subset of filter expressions: 'Col=Value'.
    Returns (col, value) or (None, None) if empty.
    Only supports simple equality and numeric values (no SQL injection surface intended).
    """
    if not filter_str:
        return None, None
    s = filter_str.strip()
    if '=' not in s:
        return None, None
    parts = s.split('=', 1)
    col = parts[0].strip()
    val = parts[1].strip()
    # Try to coerce to int
    if val.isdigit():
        return col, int(val)
    # strip quotes if present
    if (val.startswith("\"") and val.endswith("\"")) or (val.startswith("\'") and val.endswith("\'")):
        return col, val[1:-1]
    return col, val


def _get_first_value(source_table: str, source_column: str, source_filter: str):
    """Retrieve the first matching value for a mapping. Returns None if not found."""
    model = TABLE_MODEL_MAP.get(source_table.lower())
    if not model:
        return None
    col, val = _parse_simple_filter(source_filter)
    try:
        if col is None:
            # no filter: just take the first row
            obj = model.query.first()
        else:
            # only support equality for integer PK or simple string field
            obj = model.query.filter(getattr(model, col) == val).first()
        if not obj:
            return None
        # return attribute if exists
        if hasattr(obj, source_column):
            return getattr(obj, source_column)
        # else return None
        return None
    except Exception:
        return None


def preview_mappings(evaluat_id: int = None, mapping_type_id: int = None):
    """Return mappings with resolved values for previewing.
    Either specify `mapping_type_id` to preview a MappingType, or `evaluat_id` to preview mappings linked to an evaluation.
    """
    mappings = []
    evaluat = None
    if mapping_type_id:
        mt = MappingType.query.get(mapping_type_id)
        if not mt:
            return {'error': 'MappingType not found'}
        mappings = mt.sheet_mappings
    elif evaluat_id:
        evaluat = Evaluat.query.get(evaluat_id)
        if not evaluat:
            return {'error': 'Evaluation not found'}
        mappings = evaluat.sheet_mappings
    else:
        return {'error': 'mapping_type_id or evaluat_id required'}

    results = []
    for m in mappings:
        val = _get_first_value(m.Source_Table, m.Source_Column, m.Source_Filter)
        results.append({'mapping': m.To_Dict(), 'value': val})
    return {'evaluat': evaluat.Evaluat_To_Dict() if evaluat else None, 'mappings': results, 'mapping_type': mt.To_Dict() if mapping_type_id else None}


def fill_template(evaluat_id: int = None, mapping_type_id: int = None):
    """Create a filled copy of the local template using either mappings from a MappingType or mappings linked to an evaluation.
    - Loads template from Evaluat.Sheet_Local_Path
    - For each mapping, resolves value and writes it to the target cell
    - Skips writing when the existing cell has a non-empty value (preserve formulas/values)
    - Saves as a new file with suffix _filled_YYYYMMDD_HHMMSS.xlsx next to the template
    Returns dict {'success': True, 'path': path} or {'success': False, 'error': msg}
    """
    # resolve mappings
    mappings = []
    ev = None
    if mapping_type_id:
        mt = MappingType.query.get(mapping_type_id)
        if not mt:
            return {'success': False, 'error': 'MappingType not found'}
        mappings = mt.sheet_mappings
    elif evaluat_id:
        ev = Evaluat.query.get(evaluat_id)
        if not ev:
            return {'success': False, 'error': 'Evaluation not found'}
        mappings = ev.sheet_mappings
    else:
        return {'success': False, 'error': 'mapping_type_id or evaluat_id required'}

    # determine template path
    if ev:
        path = ev.Sheet_Local_Path
        if not path:
            return {'success': False, 'error': 'No template path configured for this evaluation'}
    else:
        # For MappingType-based fill, require an evaluat_id to locate a template
        return {'success': False, 'error': 'When using a MappingType, provide an evaluat_id to select the template file'}
    if not path:
        return {'success': False, 'error': 'No template path configured for this evaluation'}
    if not os.path.isabs(path):
        # allow relative to project root
        base = os.path.dirname(__file__)
        path = os.path.join(base, path)
    if not os.path.exists(path):
        return {'success': False, 'error': f'Template file not found: {path}'}

    try:
        # open two workbooks: one for writing (preserve formulas) and one with data_only
        wb = load_workbook(filename=path)
        wb_data = load_workbook(filename=path, data_only=True)
    except Exception as e:
        return {'success': False, 'error': f'Failed to open workbook: {e}'}

    modified = False
    for m in mappings:
        val = _get_first_value(m.Source_Table, m.Source_Column, m.Source_Filter)
        # target sheet
        sheet_name = m.Target_Sheet
        if sheet_name not in wb.sheetnames:
            # skip mapping if sheet not present
            continue
        ws = wb[sheet_name]
        cell_ref = m.Target_Cell
        try:
            cell = ws[cell_ref]
        except Exception:
            # invalid cell reference
            continue
        # do not overwrite if cell already contains a value (and not 0)
        if cell.value is not None and cell.value != 0 and str(cell.value).strip() != '':
            continue
        # write value (convert None to empty string)
        cell.value = '' if val is None else val
        modified = True

    # save output
    out_dir = os.path.dirname(path)
    base_name = os.path.splitext(os.path.basename(path))[0]
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_name = f"{base_name}_filled_{ts}.xlsx"
    out_path = os.path.join(out_dir, out_name)
    try:
        wb.save(out_path)
    except Exception as e:
        return {'success': False, 'error': f'Failed to save filled workbook: {e}'}

    return {'success': True, 'path': out_path, 'modified': modified}


def fill_student_sheets(evaluat_id: int = None, dry_run: bool = False, overwrite_formulas: bool = False):
    """Fill a workbook that contains one sheet per student named with numbers (1,2,3...).
    Behavior:
    - For each sheet whose name is numeric, read student name from cell B1 and match to `Studnt.Name`.
    - For rows A6..A48 read Skill.Code, find corresponding Skill and Score for given evaluat_id and student.
    - For Score.Level_Id == 1..4, write value `1` in columns E,F,G,H respectively on the same row, but only
      when the target cell is empty or contains 0 (do not overwrite other values or formulas).
    - If dry_run is True, do not save, return a summary of intended writes in `actions`.
    Returns dict with keys `success`, and either `path` (when saved) or `actions` (when dry_run).
    """
    if not evaluat_id:
        return {'success': False, 'error': 'evaluat_id required'}

    ev = Evaluat.query.get(evaluat_id)
    if not ev:
        return {'success': False, 'error': 'Evaluation not found'}

    path = ev.Sheet_Local_Path
    if not path:
        return {'success': False, 'error': 'No template path configured for this evaluation'}
    if not os.path.isabs(path):
        base = os.path.dirname(__file__)
        path = os.path.join(base, path)
    if not os.path.exists(path):
        return {'success': False, 'error': f'Template file not found: {path}'}

    try:
        wb = load_workbook(filename=path)
        wb_data = load_workbook(filename=path, data_only=True)
    except Exception as e:
        return {'success': False, 'error': f'Failed to open workbook: {e}'}

    # mapping from Level.Id to column letter
    level_to_col = {1: 'E', 2: 'F', 3: 'G', 4: 'H'}
    actions = []

    for sheet_name in wb_data.sheetnames:
        if not str(sheet_name).isdigit():
            continue
        # mark start of processing for this sheet
        actions.append({'sheet': sheet_name, 'status': 'start_processing'})
        ws_data = wb_data[sheet_name]
        ws = wb[sheet_name]
        try:
            # student name expected in B3 (use data_only workbook to get computed value)
            try:
                student_name = (ws_data['B3'].value or '')
                student_name = str(student_name).strip()
            except Exception:
                student_name = ''

            # If B1 is empty or contains a numeric placeholder (sheet number), try to resolve
            # from the formula (VLOOKUP / RECHERCHEV) or fallback to sheet index mapping.
            stud = None
            if student_name:
                stud = Studnt.query.filter(func.lower(Studnt.Name) == student_name.lower()).first()

            if not stud:
                # try to evaluate formula in the original workbook if present
                formula_val = None
                try:
                    raw = ws['B3'].value
                    if raw and isinstance(raw, str) and raw.startswith('='):
                        formula_val = _eval_vlookup_formula(raw, wb, wb_data, sheet_name)
                except Exception:
                    formula_val = None

                if formula_val:
                    student_name = str(formula_val).strip()
                    stud = Studnt.query.filter(func.lower(Studnt.Name) == student_name.lower()).first()

            if not stud:
                # as last resort, if sheet is numeric, map by index to the student's ordered list
                if str(sheet_name).isdigit():
                    try:
                        idx = int(sheet_name) - 1
                        studs = Studnt.query.filter_by(Group_Id=ev.Group_Id).order_by(Studnt.Name).all()
                        if 0 <= idx < len(studs):
                            stud = studs[idx]
                            student_name = stud.Name
                            actions.append({'sheet': sheet_name, 'status': 'mapped_by_index', 'student': student_name})
                    except Exception:
                        stud = None

            if not student_name and not stud:
                actions.append({'sheet': sheet_name, 'status': 'no_student_name'})
                continue

            if not stud:
                actions.append({'sheet': sheet_name, 'status': 'student_not_found', 'name': student_name})
                continue

            for row in range(6, 49):
                try:
                    code_val = ws_data[f'A{row}'].value
                except Exception:
                    code_val = None
                # use ws (write workbook) for writing later
                if code_val is None:
                    continue
                code_str = str(code_val).strip()
                if not code_str:
                    continue

                # case-insensitive + normalized lookup for skill code (sheet order may differ from DB)
                norm_code = _normalize_skill_code(code_str)
                try:
                    skill = Skill.query.filter(
                        func.replace(func.replace(func.upper(Skill.Code), ' ', ''), '.', '') == norm_code
                    ).first()
                except Exception:
                    try:
                        skill = Skill.query.filter(func.lower(Skill.Code) == code_str.lower()).first()
                    except Exception:
                        skill = Skill.query.filter_by(Code=code_str).first()
                if not skill:
                    actions.append({'sheet': sheet_name, 'row': row, 'skill': code_str, 'status': 'skill_not_found'})
                    continue

                score = Score.query.filter_by(Evaluat_Id=evaluat_id, Studnt_Id=stud.Id, Skill_Id=skill.Id).first()
                if not score or not getattr(score, 'Level_Id', None):
                    # no score assigned
                    continue

                level_id = int(score.Level_Id)
                col = level_to_col.get(level_id)
                if not col:
                    actions.append({'sheet': sheet_name, 'row': row, 'skill': skill.Code, 'status': 'level_out_of_range', 'level': level_id})
                    continue

                cell = ws[f"{col}{row}"]
                # prefer evaluated value from data_only workbook when deciding writability
                try:
                    existing_eval = ws_data[f"{col}{row}"].value
                except Exception:
                    existing_eval = None

                # if evaluated value is a formula string or ArrayFormula object, mark as 'unevaluated'
                unevaluated = False
                try:
                    sval = str(existing_eval)
                    if sval.startswith('='):
                        unevaluated = True
                except Exception:
                    # some openpyxl formula objects stringify differently
                    if existing_eval is not None and existing_eval.__class__.__name__.lower().endswith('arrayformula'):
                        unevaluated = True

                # if unevaluated, try to resolve simple INDIRECT patterns from the real workbook
                if unevaluated:
                    try:
                        raw_formula = ws[f"{col}{row}"].value
                        indirect_val = _eval_indirect_formula(raw_formula, wb, wb_data, sheet_name)
                        if indirect_val is not None:
                            existing_eval = indirect_val
                            unevaluated = False
                    except Exception:
                        pass

                # treat empty/0 as writable
                writable = (existing_eval is None) or (str(existing_eval).strip() == '') or (str(existing_eval).strip() == '0') or (existing_eval == 0)
                if not writable and unevaluated and overwrite_formulas:
                    writable = True

                if writable:
                    # use evaluated value for 'from' in log when available, else fall back to cell.value
                    try:
                        from_val = existing_eval
                    except Exception:
                        from_val = cell.value
                    actions.append({'sheet': sheet_name, 'row': row, 'skill': skill.Code, 'student': stud.Name, 'from': from_val, 'to': 1})
                    if not dry_run:
                        cell.value = 1
                else:
                    # log the evaluated value when possible
                    try:
                        val = existing_eval
                    except Exception:
                        val = cell.value
                    actions.append({'sheet': sheet_name, 'row': row, 'skill': skill.Code, 'student': stud.Name, 'status': 'skipped_nonzero', 'value': val})

            # end of sheet processing
            actions.append({'sheet': sheet_name, 'status': 'end_processing'})
        except Exception as e:
            actions.append({'sheet': sheet_name, 'status': 'error', 'error': str(e)})
            continue

    if dry_run:
        return {'success': True, 'dry_run': True, 'actions': actions}

    # save filled workbook next to template
    out_dir = os.path.dirname(path)
    base_name = os.path.splitext(os.path.basename(path))[0]
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_name = f"{base_name}_studentfilled_{ts}.xlsx"
    out_path = os.path.join(out_dir, out_name)
    try:
        wb.save(out_path)
    except Exception as e:
        return {'success': False, 'error': f'Failed to save filled workbook: {e}'}

    # also write JSON report next to the XLSX
    try:
        json_path = os.path.splitext(out_path)[0] + '.json'
        with open(json_path, 'w', encoding='utf-8') as jf:
            json.dump({'success': True, 'actions': actions}, jf, ensure_ascii=False, indent=2, default=str)
    except Exception as e:
        # return success but include report error
        return {'success': True, 'path': out_path, 'actions': actions, 'report_error': str(e)}

    return {'success': True, 'path': out_path, 'actions': actions, 'report_path': json_path}
